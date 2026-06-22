"""
Flask application — main entry point.
Run: python app.py
"""
from flask import Flask, request, jsonify
from flask_cors import CORS
import os, sys

# ensure modules on path
sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "modules"))

from models import init_db, get_db
from modules.auth import hash_password, verify_password, create_token, validate_token, revoke_token
from modules.ai_engine import AIEngine

app = Flask(__name__)
CORS(app, supports_credentials=True)

# ─────────────────────────────────────────────────────────────────────────────
# Auth middleware helper
# ─────────────────────────────────────────────────────────────────────────────

def _get_owner_id():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    if not token:
        return None
    return validate_token(token)


def _owner_owns_store(owner_id: int, store_id: int) -> bool:
    db = get_db()
    row = db.execute(
        "SELECT id FROM stores WHERE id = ? AND owner_id = ?", (store_id, owner_id)
    ).fetchone()
    db.close()
    return row is not None


# ─────────────────────────────────────────────────────────────────────────────
# Auth routes
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.get_json()
    email    = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    if not email or not password:
        return jsonify({"error": "Email and password required"}), 400

    db = get_db()
    owner = db.execute(
        "SELECT * FROM owners WHERE LOWER(email) = ?", (email,)
    ).fetchone()
    db.close()

    if not owner or not verify_password(password, owner["password_hash"]):
        return jsonify({"error": "Invalid credentials"}), 401

    token = create_token(owner["id"])
    db2 = get_db()
    stores = db2.execute(
        "SELECT id, name, store_type FROM stores WHERE owner_id = ?", (owner["id"],)
    ).fetchall()
    db2.close()

    return jsonify({
        "token": token,
        "owner": {"id": owner["id"], "name": owner["name"], "email": owner["email"]},
        "stores": [dict(s) for s in stores],
    })


@app.route("/api/auth/logout", methods=["POST"])
def logout():
    token = request.headers.get("Authorization", "").replace("Bearer ", "")
    revoke_token(token)
    return jsonify({"message": "Logged out"})


def _validate_password(pw: str):
    """Return an error string if the password doesn't meet policy, else None.
    Policy: min 8 chars, at least 1 uppercase, 1 lowercase, 1 digit, 1 special char.
    """
    import re
    if len(pw) < 8:
        return "Password must be at least 8 characters long."
    if not re.search(r'[A-Z]', pw):
        return "Password must contain at least one uppercase letter."
    if not re.search(r'[a-z]', pw):
        return "Password must contain at least one lowercase letter."
    if not re.search(r'[0-9]', pw):
        return "Password must contain at least one number."
    if not re.search(r'[^A-Za-z0-9]', pw):
        return "Password must contain at least one special character (e.g. @, #, !, $)."
    return None


@app.route("/api/auth/register", methods=["POST"])
def register():
    data       = request.get_json()
    name       = (data.get("name") or "").strip()
    email      = (data.get("email") or "").strip().lower()
    password   = data.get("password") or ""
    store_names= data.get("stores") or []   # list of {name, store_type}

    if not name or not email or not password:
        return jsonify({"error": "Name, email and password are required"}), 400
    if not store_names:
        return jsonify({"error": "At least one store is required"}), 400

    pw_error = _validate_password(password)
    if pw_error:
        return jsonify({"error": pw_error}), 400

    db = get_db()
    existing = db.execute("SELECT id FROM owners WHERE email = ?", (email,)).fetchone()
    if existing:
        db.close()
        return jsonify({"error": "Email already registered"}), 409

    pw_hash  = hash_password(password)
    owner_id = db.execute(
        "INSERT INTO owners (name, email, password_hash) VALUES (?,?,?)",
        (name, email, pw_hash)
    ).lastrowid
    db.commit()

    created_stores = []
    for s in store_names:
        sname = (s.get("name") or "").strip()
        stype = (s.get("store_type") or "general").strip()
        if not sname:
            continue
        sid = db.execute(
            "INSERT INTO stores (name, store_type, owner_id) VALUES (?,?,?)",
            (sname, stype, owner_id)
        ).lastrowid
        db.commit()
        # Do NOT auto-seed products; new owners start with empty inventory
        created_stores.append({"id": sid, "name": sname, "store_type": stype})

    db.close()
    token = create_token(owner_id)
    return jsonify({
        "token": token,
        "owner": {"id": owner_id, "name": name, "email": email},
        "stores": created_stores,
    }), 201


def _init_new_store(db, store_id: int, store_type: str):
    """Initialize a newly registered store with sample products."""
    from seed_data import ABC_PRODUCTS, XYZ_PRODUCTS, MNO_PRODUCTS
    import random, hashlib
    from datetime import date, timedelta
    product_map = {
        "general":     ABC_PRODUCTS[:5],
        "electronics": XYZ_PRODUCTS[:5],
        "automotive":  MNO_PRODUCTS[:5],
    }
    products = product_map.get(store_type, ABC_PRODUCTS[:5])
    for p in products:
        pid = db.execute(
            "INSERT INTO products (store_id, name, category, subcategory, brand, variant, "
            "unit_price, cost_price, is_perishable) VALUES (?,?,?,?,?,?,?,?,?)",
            (store_id, p[0], p[1], p[2], p[3], p[4], p[5], p[6], p[7])
        ).lastrowid
        db.execute(
            "INSERT OR IGNORE INTO inventory (store_id, product_id, quantity, reorder_level, max_capacity) "
            "VALUES (?,?,?,?,?)", (store_id, pid, 50, 10, 200)
        )
    db.commit()

    # Generate 3 months of sample sales
    from seed_data import _generate_sales
    cids = []
    for i in range(10):
        cid = db.execute(
            "INSERT INTO customers (store_id, name, email) VALUES (?,?,?)",
            (store_id, f"Demo_Customer_{i+1}", f"dc{i+1}@newstore.com")
        ).lastrowid
        cids.append(cid)
    db.commit()


# ─────────────────────────────────────────────────────────────────────────────
# Store routes
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/api/stores", methods=["GET"])
def list_stores():
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    db = get_db()
    stores = db.execute(
        "SELECT id, name, store_type FROM stores WHERE owner_id = ?", (owner_id,)
    ).fetchall()
    db.close()
    return jsonify([dict(s) for s in stores])


@app.route("/api/stores", methods=["POST"])
def add_store():
    """Allow authenticated owners to add a new store."""
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    data   = request.get_json()
    sname  = (data.get("name") or "").strip()
    stype  = (data.get("store_type") or "general").strip()
    if not sname:
        return jsonify({"error": "Store name required"}), 400
    db = get_db()
    sid = db.execute(
        "INSERT INTO stores (name, store_type, owner_id) VALUES (?,?,?)",
        (sname, stype, owner_id)
    ).lastrowid
    db.commit()
    _init_new_store(db, sid, stype)
    db.close()
    return jsonify({"id": sid, "name": sname, "store_type": stype}), 201


# ─────────────────────────────────────────────────────────────────────────────
# Dashboard
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/api/stores/<int:store_id>/dashboard", methods=["GET"])
def dashboard(store_id):
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    if not _owner_owns_store(owner_id, store_id):
        return jsonify({"error": "Forbidden"}), 403
    engine = AIEngine(store_id)
    try:
        summary = engine.dashboard_summary()
    finally:
        engine.close()
    return jsonify(summary)


# ─────────────────────────────────────────────────────────────────────────────
# Inventory
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/api/stores/<int:store_id>/inventory", methods=["GET"])
def inventory(store_id):
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    if not _owner_owns_store(owner_id, store_id):
        return jsonify({"error": "Forbidden"}), 403
    engine = AIEngine(store_id)
    try:
        inv = engine.inventory_status()
    finally:
        engine.close()
    return jsonify(inv)


# ─────────────────────────────────────────────────────────────────────────────
# AI Insights / Recommendations
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/api/stores/<int:store_id>/insights", methods=["GET"])
def insights(store_id):
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    if not _owner_owns_store(owner_id, store_id):
        return jsonify({"error": "Forbidden"}), 403
    engine = AIEngine(store_id)
    try:
        recs = engine.generate_recommendations()
    finally:
        engine.close()
    return jsonify(recs)


@app.route("/api/owner/all-recommendations", methods=["GET"])
def all_recommendations():
    """Combined recommendations for all stores owned by this owner."""
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    db = get_db()
    stores = db.execute(
        "SELECT id, name, store_type FROM stores WHERE owner_id = ?", (owner_id,)
    ).fetchall()
    db.close()
    all_recs = []
    for store in stores:
        engine = AIEngine(store["id"])
        try:
            recs = engine.generate_recommendations()
        finally:
            engine.close()
        for r in recs:
            r["store_id"]   = store["id"]
            r["store_name"] = store["name"]
            r["store_type"] = store["store_type"]
        all_recs.extend(recs)
    # Sort combined: critical first
    priority_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    all_recs.sort(key=lambda x: priority_order.get(x.get("priority", "low"), 3))
    return jsonify(all_recs)


# ─────────────────────────────────────────────────────────────────────────────
# Sales data for charts
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/api/stores/<int:store_id>/sales/monthly", methods=["GET"])
def sales_monthly(store_id):
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    if not _owner_owns_store(owner_id, store_id):
        return jsonify({"error": "Forbidden"}), 403
    engine = AIEngine(store_id)
    try:
        data = engine.monthly_sales()
    finally:
        engine.close()
    return jsonify(data)


@app.route("/api/stores/<int:store_id>/sales/category", methods=["GET"])
def sales_category(store_id):
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    if not _owner_owns_store(owner_id, store_id):
        return jsonify({"error": "Forbidden"}), 403
    engine = AIEngine(store_id)
    try:
        data = engine.category_sales()
    finally:
        engine.close()
    return jsonify(data)


@app.route("/api/stores/<int:store_id>/sales/stockout-losses", methods=["GET"])
def stockout_losses(store_id):
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    if not _owner_owns_store(owner_id, store_id):
        return jsonify({"error": "Forbidden"}), 403
    engine = AIEngine(store_id)
    try:
        data = engine.stockout_losses()
    finally:
        engine.close()
    return jsonify(data)


@app.route("/api/stores/<int:store_id>/sales/gst-summary", methods=["GET"])
def gst_summary(store_id):
    """GST paid per Financial Year (Apr–Mar) for the store.
    Uses 18% GST on revenue as a simplified flat rate for demo purposes.
    Returns last 2 complete FYs + current FY YTD.
    """
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    if not _owner_owns_store(owner_id, store_id):
        return jsonify({"error": "Forbidden"}), 403

    from datetime import date
    GST_RATE = 0.18  # 18% flat for demo

    db = get_db()
    # Pull monthly revenue for this store
    rows = db.execute("""
        SELECT strftime('%Y', order_date) AS yr,
               strftime('%m', order_date) AS mo,
               SUM(total_amount) AS revenue
        FROM orders
        WHERE store_id = ? AND status = 'completed'
        GROUP BY yr, mo
        ORDER BY yr, mo
    """, (store_id,)).fetchall()
    db.close()

    # Group into Financial Years (Apr y -> Mar y+1  labelled "FY y/y+1")
    fy_revenue = {}
    for r in rows:
        yr = int(r["yr"]); mo = int(r["mo"])
        # FY key: the April-start year
        fy_start = yr if mo >= 4 else yr - 1
        key = f"FY {fy_start}-{str(fy_start+1)[-2:]}"
        fy_revenue[key] = fy_revenue.get(key, 0.0) + r["revenue"]

    today = date.today()
    cur_fy_start = today.year if today.month >= 4 else today.year - 1
    cur_fy_key   = f"FY {cur_fy_start}-{str(cur_fy_start+1)[-2:]}"

    # Build result — last 2 completed FYs + current FY YTD
    result = []
    for fy_key, rev in sorted(fy_revenue.items()):
        is_current = (fy_key == cur_fy_key)
        result.append({
            "fy":         fy_key,
            "revenue":    round(rev, 2),
            "gst_paid":   round(rev * GST_RATE, 2),
            "is_current": is_current,
            "ytd":        is_current,
        })

    # Return only last 2 complete FYs + current FY
    complete = [r for r in result if not r["is_current"]]
    current  = [r for r in result if r["is_current"]]
    out = complete[-2:] + current
    return jsonify(out)


# ─────────────────────────────────────────────────────────────────────────────
# Products CRUD
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/api/stores/<int:store_id>/products", methods=["GET"])
def list_products(store_id):
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    if not _owner_owns_store(owner_id, store_id):
        return jsonify({"error": "Forbidden"}), 403
    db = get_db()
    rows = db.execute("""
        SELECT p.id, p.name, p.category, p.subcategory, p.brand, p.variant,
               p.unit_price, p.cost_price, p.is_perishable,
               COALESCE(i.quantity, 0) AS quantity,
               COALESCE(i.reorder_level, 10) AS reorder_level,
               COALESCE(i.max_capacity, 200) AS max_capacity
        FROM products p
        LEFT JOIN inventory i ON i.product_id = p.id AND i.store_id = p.store_id
        WHERE p.store_id = ?
        ORDER BY p.category, p.name
    """, (store_id,)).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/stores/<int:store_id>/products", methods=["POST"])
def add_product(store_id):
    """Add a new product and create its inventory row."""
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    if not _owner_owns_store(owner_id, store_id):
        return jsonify({"error": "Forbidden"}), 403

    data = request.get_json()
    name         = (data.get("name") or "").strip()
    category     = (data.get("category") or "general").strip()
    subcategory  = (data.get("subcategory") or "").strip()
    brand        = (data.get("brand") or "").strip()
    variant      = (data.get("variant") or "").strip()
    unit_price   = float(data.get("unit_price") or 0)
    cost_price   = float(data.get("cost_price") or 0)
    is_perishable= int(bool(data.get("is_perishable", False)))
    quantity     = int(data.get("quantity") or 0)
    reorder_level= int(data.get("reorder_level") or 10)
    max_capacity = int(data.get("max_capacity") or 200)

    if not name:
        return jsonify({"error": "Product name is required"}), 400
    if unit_price <= 0:
        return jsonify({"error": "Unit price must be > 0"}), 400

    db = get_db()
    pid = db.execute(
        "INSERT INTO products (store_id, name, category, subcategory, brand, variant, "
        "unit_price, cost_price, is_perishable) VALUES (?,?,?,?,?,?,?,?,?)",
        (store_id, name, category, subcategory, brand, variant,
         unit_price, cost_price, is_perishable)
    ).lastrowid
    db.execute(
        "INSERT OR REPLACE INTO inventory (store_id, product_id, quantity, reorder_level, max_capacity) "
        "VALUES (?,?,?,?,?)",
        (store_id, pid, quantity, reorder_level, max_capacity)
    )
    db.commit()
    db.close()
    return jsonify({"id": pid, "name": name, "category": category, "quantity": quantity}), 201


@app.route("/api/stores/<int:store_id>/products/<int:product_id>", methods=["PUT"])
def update_product(store_id, product_id):
    """Update inventory quantity / reorder levels for a product."""
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    if not _owner_owns_store(owner_id, store_id):
        return jsonify({"error": "Forbidden"}), 403

    data = request.get_json()
    db   = get_db()
    # Verify product belongs to store
    prod = db.execute("SELECT id FROM products WHERE id=? AND store_id=?",
                      (product_id, store_id)).fetchone()
    if not prod:
        db.close()
        return jsonify({"error": "Product not found"}), 404

    # Update inventory fields if provided
    if "quantity" in data:
        db.execute("UPDATE inventory SET quantity=?, last_updated=CURRENT_TIMESTAMP "
                   "WHERE store_id=? AND product_id=?",
                   (int(data["quantity"]), store_id, product_id))
    if "reorder_level" in data:
        db.execute("UPDATE inventory SET reorder_level=? WHERE store_id=? AND product_id=?",
                   (int(data["reorder_level"]), store_id, product_id))
    if "max_capacity" in data:
        db.execute("UPDATE inventory SET max_capacity=? WHERE store_id=? AND product_id=?",
                   (int(data["max_capacity"]), store_id, product_id))
    # Update product fields if provided
    if "unit_price" in data:
        db.execute("UPDATE products SET unit_price=? WHERE id=?",
                   (float(data["unit_price"]), product_id))
    if "cost_price" in data:
        db.execute("UPDATE products SET cost_price=? WHERE id=?",
                   (float(data["cost_price"]), product_id))
    db.commit()
    db.close()
    return jsonify({"message": "Updated"})


@app.route("/api/stores/<int:store_id>/products/<int:product_id>", methods=["DELETE"])
def delete_product(store_id, product_id):
    """Delete a product and its inventory row from the store."""
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    if not _owner_owns_store(owner_id, store_id):
        return jsonify({"error": "Forbidden"}), 403

    db = get_db()
    prod = db.execute("SELECT id FROM products WHERE id=? AND store_id=?",
                      (product_id, store_id)).fetchone()
    if not prod:
        db.close()
        return jsonify({"error": "Product not found"}), 404

    db.execute("DELETE FROM order_items WHERE product_id=?", (product_id,))
    db.execute("DELETE FROM inventory WHERE product_id=? AND store_id=?", (product_id, store_id))
    db.execute("DELETE FROM products WHERE id=? AND store_id=?", (product_id, store_id))
    db.commit()
    db.close()
    return jsonify({"message": "Deleted"})


# ─────────────────────────────────────────────────────────────────────────────
# Excel template download + historical data upload
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/api/stores/<int:store_id>/upload/template", methods=["GET"])
def download_template(store_id):
    """Generate and return an Excel template for historical data upload."""
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    if not _owner_owns_store(owner_id, store_id):
        return jsonify({"error": "Forbidden"}), 403

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        from flask import send_file
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    wb = openpyxl.Workbook()

    # ── Sheet 1: Products ────────────────────────────────────────────────────
    ws_prod = wb.active
    ws_prod.title = "Products"

    RED    = Font(color="FF0000", bold=True)
    BOLD   = Font(bold=True)
    HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
    HDR_FONT = Font(color="FFFFFF", bold=True)
    REQ_FILL = PatternFill("solid", fgColor="FFF3CD")
    THIN = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    prod_headers = [
        ("Product Name *",    True,  "Name of the product (e.g. Football Size 5)"),
        ("Category *",        True,  "Main category (e.g. sports, electronics, vehicles)"),
        ("Subcategory",       False, "Optional sub-group (e.g. team sports, racket sports)"),
        ("Brand",             False, "Brand name (e.g. Nike, Adidas)"),
        ("Variant",           False, "Size/model/variant (e.g. Size 5, 200g)"),
        ("Unit Price (₹) *",  True,  "Selling price per unit"),
        ("Cost Price (₹) *",  True,  "Purchase/cost price per unit"),
        ("Current Stock *",   True,  "Current quantity in stock"),
        ("Reorder Level",     False, "Stock level at which reorder is triggered (default 10)"),
        ("Max Capacity",      False, "Maximum storage capacity (default 200)"),
        ("Is Perishable",     False, "Enter YES if perishable, otherwise leave blank"),
    ]

    for col, (header, required, hint) in enumerate(prod_headers, 1):
        cell = ws_prod.cell(row=1, column=col, value=header)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN
        ws_prod.cell(row=2, column=col, value=f"[{hint}]").font = Font(italic=True, color="888888", size=8)
        ws_prod.cell(row=2, column=col).border = THIN
        ws_prod.column_dimensions[get_column_letter(col)].width = max(18, len(header) + 2)

    # Mark required columns with bold white font (already set via HDR_FONT; keep consistent)
    for col, (header, required, _) in enumerate(prod_headers, 1):
        if required:
            cell = ws_prod.cell(row=1, column=col)
            cell.font = Font(color="FFFFFF", bold=True)

    # ONE example row only
    samples = [
        ["Football Size 5", "sports", "team sports", "Nike", "Size 5", 799, 550, 50, 10, 200, ""],
    ]
    for r, row in enumerate(samples, 3):
        for c, val in enumerate(row, 1):
            cell = ws_prod.cell(row=r, column=c, value=val)
            cell.border = THIN
            if prod_headers[c-1][1]:  # required column
                cell.fill = PatternFill("solid", fgColor="EBF5FF")

    ws_prod.row_dimensions[1].height = 30
    ws_prod.row_dimensions[2].height = 24
    ws_prod.freeze_panes = "A3"

    # ── Sheet 2: Sales History ───────────────────────────────────────────────
    ws_sales = wb.create_sheet("Sales History")

    sales_headers = [
        ("Order Date *",      True,  "Date of sale in YYYY-MM-DD format (e.g. 2024-06-15)"),
        ("Product Name *",    True,  "Must match a product name from Products sheet"),
        ("Quantity Sold *",   True,  "Number of units sold in this transaction"),
        ("Unit Price (₹) *",  True,  "Selling price per unit at time of sale"),
        ("Customer Name",     False, "Optional — customer name or ID"),
    ]

    for col, (header, required, hint) in enumerate(sales_headers, 1):
        cell = ws_sales.cell(row=1, column=col, value=header)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN
        ws_sales.cell(row=2, column=col, value=f"[{hint}]").font = Font(italic=True, color="888888", size=8)
        ws_sales.cell(row=2, column=col).border = THIN
        ws_sales.column_dimensions[get_column_letter(col)].width = max(20, len(header) + 2)

    # ONE example row only
    sales_samples = [
        ["2024-06-01", "Football Size 5", 3, 799, "Walk-in Customer"],
    ]
    for r, row in enumerate(sales_samples, 3):
        for c, val in enumerate(row, 1):
            cell = ws_sales.cell(row=r, column=c, value=val)
            cell.border = THIN
            if sales_headers[c-1][1]:
                cell.fill = PatternFill("solid", fgColor="EBF5FF")

    ws_sales.row_dimensions[1].height = 30
    ws_sales.row_dimensions[2].height = 24
    ws_sales.freeze_panes = "A3"

    # ── Sheet 3: Instructions ────────────────────────────────────────────────
    ws_info = wb.create_sheet("Instructions")
    instructions = [
        ("RetailAI — Historical Data Upload Template", True),
        ("", False),
        ("INSTRUCTIONS", True),
        ("1. Fill the 'Products' sheet first with all your products.", False),
        ("2. Fill the 'Sales History' sheet with your historical sales data.", False),
        ("3. Columns marked with * (red star) are REQUIRED — do not leave them blank.", False),
        ("4. Date format must be YYYY-MM-DD (e.g. 2024-06-15).", False),
        ("5. Product Name in Sales History must exactly match the name in Products sheet.", False),
        ("6. Delete the sample rows (rows 3 onwards) before uploading your data.", False),
        ("7. Do not modify column headers in row 1.", False),
        ("", False),
        ("REQUIRED FIELDS (marked with * in headers):", True),
        ("  Products sheet:      Product Name, Category, Unit Price, Cost Price, Current Stock", False),
        ("  Sales History sheet: Order Date, Product Name, Quantity Sold, Unit Price", False),
        ("", False),
        ("NOTES:", True),
        ("  - Is Perishable: enter YES for dairy, fresh produce, frozen goods, etc.", False),
        ("  - Reorder Level: if left blank, defaults to 10 units", False),
        ("  - Max Capacity: if left blank, defaults to 200 units", False),
        ("  - Rows 1-2 are headers and hints — your data starts from row 3", False),
    ]
    for r, (text, bold) in enumerate(instructions, 1):
        cell = ws_info.cell(row=r, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=12 if r == 1 else 11)
        ws_info.column_dimensions["A"].width = 80

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="RetailAI Template.xlsx"
    )


@app.route("/api/stores/<int:store_id>/upload/data", methods=["POST"])
def upload_excel_data(store_id):
    """Parse uploaded Excel file and insert products + sales history."""
    owner_id = _get_owner_id()
    if not owner_id:
        return jsonify({"error": "Unauthorized"}), 401
    if not _owner_owns_store(owner_id, store_id):
        return jsonify({"error": "Forbidden"}), 403

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        return jsonify({"error": "Only .xlsx files are accepted"}), 400

    try:
        import io as _io
        wb = openpyxl.load_workbook(_io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return jsonify({"error": f"Could not read Excel file: {e}"}), 400

    db = get_db()
    report = {"products_added": 0, "products_skipped": 0, "sales_added": 0,
              "sales_skipped": 0, "errors": []}

    # ── Parse Products sheet ─────────────────────────────────────────────────
    if "Products" in wb.sheetnames:
        ws = wb["Products"]
        product_name_to_id = {}
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            name        = str(row[0]).strip()
            category    = str(row[1]).strip() if row[1] else "general"
            subcategory = str(row[2]).strip() if row[2] else ""
            brand       = str(row[3]).strip() if row[3] else ""
            variant     = str(row[4]).strip() if row[4] else ""
            try:
                unit_price   = float(row[5]) if row[5] else None
                cost_price   = float(row[6]) if row[6] else None
                quantity     = int(float(row[7])) if row[7] is not None else 0
                reorder_lvl  = int(float(row[8])) if row[8] is not None else 10
                max_cap      = int(float(row[9])) if row[9] is not None else 200
                is_perishable= 1 if str(row[10] or "").strip().upper() == "YES" else 0
            except (ValueError, TypeError) as e:
                report["errors"].append(f"Product '{name}': invalid numeric value — {e}")
                report["products_skipped"] += 1
                continue

            if not name or unit_price is None or cost_price is None:
                report["errors"].append(f"Row skipped: missing required field (name/unit_price/cost_price)")
                report["products_skipped"] += 1
                continue

            # Check if product already exists by name for this store
            existing = db.execute(
                "SELECT id FROM products WHERE store_id=? AND LOWER(name)=LOWER(?)",
                (store_id, name)
            ).fetchone()
            if existing:
                pid = existing["id"]
                # Update inventory quantity
                db.execute("UPDATE inventory SET quantity=? WHERE store_id=? AND product_id=?",
                           (quantity, store_id, pid))
                product_name_to_id[name.lower()] = pid
                report["products_skipped"] += 1  # already existed — updated qty
            else:
                pid = db.execute(
                    "INSERT INTO products (store_id, name, category, subcategory, brand, variant, "
                    "unit_price, cost_price, is_perishable) VALUES (?,?,?,?,?,?,?,?,?)",
                    (store_id, name, category, subcategory, brand, variant,
                     unit_price, cost_price, is_perishable)
                ).lastrowid
                db.execute(
                    "INSERT OR IGNORE INTO inventory (store_id, product_id, quantity, reorder_level, max_capacity) "
                    "VALUES (?,?,?,?,?)",
                    (store_id, pid, quantity, reorder_lvl, max_cap)
                )
                product_name_to_id[name.lower()] = pid
                report["products_added"] += 1
        db.commit()

    # ── Parse Sales History sheet ────────────────────────────────────────────
    if "Sales History" in wb.sheetnames:
        ws_s = wb["Sales History"]
        # Build product name map if not already populated
        if not product_name_to_id:
            for prow in db.execute("SELECT id, name FROM products WHERE store_id=?", (store_id,)).fetchall():
                product_name_to_id[prow["name"].lower()] = prow["id"]

        # Ensure at least one customer exists
        cust = db.execute("SELECT id FROM customers WHERE store_id=? LIMIT 1", (store_id,)).fetchone()
        if not cust:
            cust_id = db.execute(
                "INSERT INTO customers (store_id, name, email) VALUES (?,?,?)",
                (store_id, "Uploaded Customer", "upload@import.com")
            ).lastrowid
            db.commit()
        else:
            cust_id = cust["id"]

        for row in ws_s.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            order_date_raw = row[0]
            product_name   = str(row[1]).strip() if row[1] else ""
            try:
                qty        = int(float(row[2])) if row[2] else None
                unit_price = float(row[3]) if row[3] else None
            except (ValueError, TypeError):
                report["sales_skipped"] += 1
                report["errors"].append(f"Sale row: invalid qty/price for '{product_name}'")
                continue

            if not order_date_raw or not product_name or qty is None or unit_price is None:
                report["sales_skipped"] += 1
                continue

            # Normalise date
            from datetime import date as _date
            if hasattr(order_date_raw, 'strftime'):
                order_date = order_date_raw.strftime("%Y-%m-%d")
            else:
                try:
                    order_date = str(order_date_raw).strip()[:10]
                    _date.fromisoformat(order_date)  # validate
                except ValueError:
                    report["sales_skipped"] += 1
                    report["errors"].append(f"Sale row: bad date '{order_date_raw}'")
                    continue

            pid = product_name_to_id.get(product_name.lower())
            if not pid:
                report["sales_skipped"] += 1
                report["errors"].append(f"Sale row: product '{product_name}' not found in Products sheet or DB")
                continue

            total = qty * unit_price
            oid   = db.execute(
                "INSERT INTO orders (store_id, customer_id, order_date, total_amount, status) "
                "VALUES (?,?,?,?,'completed')",
                (store_id, cust_id, order_date, round(total, 2))
            ).lastrowid
            db.execute(
                "INSERT INTO order_items (order_id, product_id, quantity, unit_price) VALUES (?,?,?,?)",
                (oid, pid, qty, unit_price)
            )
            report["sales_added"] += 1

        db.commit()

    db.close()
    return jsonify({
        "message": "Upload complete",
        "products_added":   report["products_added"],
        "products_updated": report["products_skipped"],
        "sales_added":      report["sales_added"],
        "sales_skipped":    report["sales_skipped"],
        "errors":           report["errors"][:20],  # cap error list
    })


# ─────────────────────────────────────────────────────────────────────────────
# Health check
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "message": "Retail AI App backend running"})


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Initializing database...")
    init_db()
    print("Starting Flask server on http://localhost:5000")
    app.run(debug=False, port=5000, host="0.0.0.0")
